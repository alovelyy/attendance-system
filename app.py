import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, time, date
import io
import zipfile
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re

# ============================================================
# CONFIGURATION
# ============================================================
WEEKEND_DAY = 4  # Friday
NORMAL_HOURS_PERMANENT = 8
NORMAL_HOURS_TEMPORARY = 10
RAMADAN_NORMAL_HOURS = 6

# ============================================================
# FUTURISTIC LOGIN SCREEN (NEW DESIGN)
# ============================================================
def render_login():
    st.markdown("""
    <style>
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        header {visibility: hidden;}
        .stApp {margin: 0; padding: 0; background: #0a0e1a;}
        .block-container {padding: 0 !important; max-width: 100% !important;}
        .stApp > div:first-child {padding: 0 !important;}
        .st-emotion-cache-1r6slb0 {padding: 0 !important;}
        .st-emotion-cache-1gv3huu {padding: 0 !important;}
        body {background: #0a0e1a; margin: 0; overflow: hidden; height: 100vh; width: 100vw;}
        
        /* ---- ANIMATED MESH BACKGROUND ---- */
        #mesh-bg {
            position: fixed;
            top: 0; left: 0;
            width: 100%; height: 100%;
            z-index: 0;
            background: radial-gradient(ellipse at 20% 30%, #1a2a4a, #0a0e1a 80%);
            overflow: hidden;
        }
        .mesh-orb {
            position: absolute;
            border-radius: 50%;
            filter: blur(100px);
            opacity: 0.4;
            animation: floatOrb 15s infinite alternate ease-in-out;
        }
        .mesh-orb:nth-child(1) {
            width: 600px; height: 600px;
            top: -10%; left: -10%;
            background: #f0c040;
        }
        .mesh-orb:nth-child(2) {
            width: 500px; height: 500px;
            bottom: -10%; right: -10%;
            background: #00ccff;
            animation-delay: 6s;
        }
        .mesh-orb:nth-child(3) {
            width: 400px; height: 400px;
            top: 40%; left: 50%;
            background: #ff6b6b;
            animation-delay: 10s;
        }
        @keyframes floatOrb {
            0% { transform: translate(0, 0) scale(1); }
            100% { transform: translate(60px, 40px) scale(1.1); }
        }
        .particle {
            position: absolute;
            width: 3px; height: 3px;
            background: rgba(255,255,255,0.4);
            border-radius: 50%;
            animation: twinkle 4s infinite alternate;
        }
        @keyframes twinkle {
            0% { opacity: 0.1; transform: scale(0.8); }
            100% { opacity: 0.8; transform: scale(1.2); }
        }
        
        /* ---- FLEX CONTAINER ---- */
        .login-container {
            position: fixed;
            top: 0; left: 0;
            width: 100%; height: 100%;
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            z-index: 5;
            pointer-events: none;
        }
        .login-container > * { pointer-events: auto; }
        
        /* ---- TITLE SECTION (Neon DPSR) ---- */
        .title-section {
            text-align: center;
            margin-bottom: 30px;
            opacity: 0;
            animation: fadeUp 1.8s ease-out forwards;
            animation-delay: 0.5s;
        }
        .neon-dpsr {
            font-size: clamp(60px, 12vw, 110px);
            font-weight: 900;
            color: #f0c040;
            text-shadow: 0 0 30px rgba(240,192,64,0.3), 0 0 60px rgba(240,192,64,0.1);
            letter-spacing: 0.2em;
            font-family: 'Impact', 'Arial Black', sans-serif;
            animation: flicker 3s infinite alternate;
        }
        @keyframes flicker {
            0% { text-shadow: 0 0 30px rgba(240,192,64,0.3), 0 0 60px rgba(240,192,64,0.1); }
            100% { text-shadow: 0 0 40px rgba(240,192,64,0.6), 0 0 80px rgba(240,192,64,0.2); }
        }
        .company-name {
            font-size: clamp(24px, 4vw, 48px);
            font-weight: 700;
            color: #fff;
            letter-spacing: 0.15em;
            text-shadow: 0 0 20px rgba(0,200,255,0.2);
            margin-top: -6px;
        }
        .company-sub {
            font-size: clamp(14px, 1.8vw, 20px);
            color: rgba(255,255,255,0.3);
            letter-spacing: 0.5em;
            font-weight: 300;
            margin-top: 4px;
        }
        .title-divider {
            width: 120px; height: 2px;
            background: linear-gradient(90deg, transparent, #f0c040, transparent);
            margin: 10px auto 0;
        }
        @keyframes fadeUp {
            0% { opacity: 0; transform: translateY(30px); }
            100% { opacity: 1; transform: translateY(0); }
        }
        
        /* ---- LOGIN CARD (Glass + Glow Border) ---- */
        .login-wrapper {
            width: 90%;
            max-width: 400px;
            opacity: 0;
            animation: cardRise 1.4s ease-out forwards;
            animation-delay: 1.8s;
        }
        @keyframes cardRise {
            0% { opacity: 0; transform: translateY(40px) scale(0.95); }
            100% { opacity: 1; transform: translateY(0) scale(1); }
        }
        .login-card {
            padding: 30px 30px 25px;
            background: rgba(10, 14, 26, 0.6);
            backdrop-filter: blur(20px);
            border: 1px solid rgba(255,215,0,0.15);
            border-radius: 28px;
            box-shadow: 0 30px 80px rgba(0,0,0,0.8), 0 0 40px rgba(240,192,64,0.03);
            position: relative;
            overflow: hidden;
        }
        .login-card::before {
            content: '';
            position: absolute;
            inset: -2px;
            border-radius: 30px;
            background: conic-gradient(from 0deg, transparent, rgba(240,192,64,0.2), transparent 60%);
            animation: spinGlow 8s linear infinite;
            z-index: -1;
            mask: linear-gradient(#fff 0 0) content-box, linear-gradient(#fff 0 0);
            -webkit-mask: linear-gradient(#fff 0 0) content-box, linear-gradient(#fff 0 0);
            mask-composite: exclude;
            -webkit-mask-composite: xor;
            padding: 2px;
        }
        @keyframes spinGlow {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        /* Streamlit inputs override */
        .stTextInput > div > div > input {
            background: rgba(255,255,255,0.05) !important;
            border: 1px solid rgba(255,255,255,0.08) !important;
            border-radius: 50px !important;
            color: #fff !important;
            padding: 14px 20px !important;
            font-size: 16px !important;
            width: 100% !important;
            transition: all 0.3s !important;
            position: relative;
            z-index: 1;
        }
        .stTextInput > div > div > input:focus {
            border-color: #f0c040 !important;
            box-shadow: 0 0 30px rgba(240,192,64,0.15), inset 0 0 20px rgba(240,192,64,0.05) !important;
        }
        .stButton > button {
            width: 100% !important;
            background: linear-gradient(135deg, #f0c040, #d4a020) !important;
            border: none !important;
            border-radius: 50px !important;
            color: #0b0e17 !important;
            font-size: 18px !important;
            font-weight: 700 !important;
            padding: 14px !important;
            margin-top: 16px !important;
            box-shadow: 0 8px 30px rgba(240,192,64,0.25) !important;
            transition: all 0.3s !important;
            position: relative;
            z-index: 1;
            animation: pulseBtn 2s infinite ease-in-out;
        }
        @keyframes pulseBtn {
            0% { transform: scale(1); }
            50% { transform: scale(1.02); }
            100% { transform: scale(1); }
        }
        .stButton > button:hover {
            transform: scale(1.03) !important;
            box-shadow: 0 12px 40px rgba(240,192,64,0.5) !important;
        }
        .stAlert {
            background: rgba(255,107,107,0.08) !important;
            border: 1px solid rgba(255,107,107,0.15) !important;
            border-radius: 12px !important;
            color: #ff6b6b !important;
            padding: 10px 20px !important;
            margin-top: 12px !important;
            font-size: 14px !important;
            position: relative;
            z-index: 1;
        }
        
        /* ---- CONTACT CARD (Minimal) ---- */
        .contact-wrapper {
            width: 90%;
            max-width: 400px;
            margin-top: 20px;
            opacity: 0;
            animation: cardRise 1.4s ease-out forwards;
            animation-delay: 2.2s;
        }
        .contact-card {
            padding: 12px 20px 12px;
            background: rgba(10, 14, 26, 0.4);
            backdrop-filter: blur(12px);
            border: 1px solid rgba(255,255,255,0.04);
            border-radius: 16px;
            text-align: left;
        }
        .contact-info {
            color: rgba(255,255,255,0.5);
            font-size: 14px;
            line-height: 1.7;
            font-weight: 300;
            letter-spacing: 0.3px;
        }
        .contact-info .name {
            color: #ffffff;
            font-weight: 500;
            font-size: 16px;
        }
        .contact-info .detail span {
            display: inline-block;
            margin-right: 8px;
            opacity: 0.4;
        }
        
        /* Responsive */
        @media (max-height: 700px) {
            .neon-dpsr { font-size: clamp(40px, 8vw, 60px); }
            .company-name { font-size: clamp(20px, 3vw, 32px); }
            .company-sub { font-size: 12px; }
            .login-card { padding: 20px 20px 16px; }
            .stTextInput > div > div > input { padding: 12px 16px; font-size: 14px; }
            .stButton > button { padding: 12px; font-size: 16px; }
            .contact-card { padding: 10px 16px 10px; }
            .contact-info { font-size: 13px; }
            .contact-info .name { font-size: 14px; }
        }
        @media (max-height: 600px) {
            .neon-dpsr { font-size: clamp(30px, 6vw, 45px); }
            .company-name { font-size: clamp(16px, 2.5vw, 24px); }
            .title-section { margin-bottom: 15px; }
            .login-card { padding: 14px 16px 12px; }
            .contact-wrapper { margin-top: 10px; }
        }
    </style>
    """, unsafe_allow_html=True)

    # ---- Background: mesh orbs + particles ----
    st.markdown("""
    <div id="mesh-bg">
        <div class="mesh-orb"></div>
        <div class="mesh-orb"></div>
        <div class="mesh-orb"></div>
    </div>
    <script>
        (function() {
            const bg = document.getElementById('mesh-bg');
            for (let i=0; i<120; i++) {
                const p = document.createElement('div');
                p.className = 'particle';
                p.style.left = Math.random() * 100 + '%';
                p.style.top = Math.random() * 100 + '%';
                p.style.width = (1 + Math.random() * 3) + 'px';
                p.style.height = p.style.width;
                p.style.animationDelay = Math.random() * 5 + 's';
                p.style.animationDuration = (3 + Math.random() * 3) + 's';
                bg.appendChild(p);
            }
        })();
    </script>
    """, unsafe_allow_html=True)

    # ---- Title Section ----
    st.markdown("""
    <div class="login-container">
        <div class="title-section">
            <div class="neon-dpsr">DPSR</div>
            <div class="company-name">Doosan Power Systems Arabia</div>
            <div class="company-sub">— ULTIMATE ATTENDANCE SYSTEM —</div>
            <div class="title-divider"></div>
        </div>
    """, unsafe_allow_html=True)

    # ---- Login Card ----
    st.markdown("""
        <div class="login-wrapper">
            <div class="login-card">
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        username = st.text_input("Username", placeholder="👤 Username", key="login_username", label_visibility="collapsed")
        password = st.text_input("Password", placeholder="🔑 Password", type="password", key="login_password", label_visibility="collapsed")
        if st.button("🚀 Login", use_container_width=True):
            if username == "AhmedShawky" and password == "iloveshawky":
                st.session_state.logged_in = True
                st.rerun()
            else:
                st.error("❌ Invalid username or password.")

    st.markdown("""
            </div>
        </div>
    """, unsafe_allow_html=True)

    # ---- Contact Card ----
    st.markdown("""
        <div class="contact-wrapper">
            <div class="contact-card">
                <div class="contact-info">
                    <div class="name">Ahmed Shawky</div>
                    <div class="detail"><span>📞</span> +201095214911</div>
                    <div class="detail"><span>✉️</span> ahmedshawkyqz@gmail.com</div>
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)


# ============================================================
# MODE SELECTION
# ============================================================
def render_mode_selection():
    st.title("📊 Choose Your Attendance System")
    st.markdown("Select the type of data you want to process:")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("""
        <div style="background:rgba(255,255,255,0.05);border-radius:20px;padding:30px;text-align:center;border:1px solid rgba(255,255,255,0.1);">
            <h2 style="color:#f0c040;">🖥️ Machine</h2>
            <p style="color:rgba(255,255,255,0.6);">Raw logs with ID and DATE</p>
            <p style="font-size:12px;color:rgba(255,255,255,0.3);">Standard fingerprint/access logs</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Select Machine", key="btn_machine", use_container_width=True):
            st.session_state.app_mode = "machine"
            st.rerun()

    with col2:
        st.markdown("""
        <div style="background:rgba(255,255,255,0.05);border-radius:20px;padding:30px;text-align:center;border:1px solid rgba(255,255,255,0.1);">
            <h2 style="color:#f0c040;">🧬 Bio</h2>
            <p style="color:rgba(255,255,255,0.6);">Monthly attendance matrix</p>
            <p style="font-size:12px;color:rgba(255,255,255,0.3);">Excel with Employee ID, Name, Dates</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Select Bio", key="btn_bio", use_container_width=True):
            st.session_state.app_mode = "bio"
            st.rerun()


# ============================================================
# MACHINE MODE FUNCTIONS
# ============================================================
def process_logs(logs_df):
    logs_df['DATE'] = pd.to_datetime(logs_df['DATE'])
    records = []
    for emp_id, group in logs_df.groupby('ID'):
        for dt, day_group in group.groupby(group['DATE'].dt.date):
            times = day_group['DATE'].tolist()
            if not times:
                continue
            times.sort()
            earliest = times[0]
            latest = times[-1]
            gross_sec = (latest - earliest).total_seconds()
            gross_hours = gross_sec / 3600.0
            records.append({
                'EmployeeID': str(emp_id).strip(),
                'Date': dt,
                'CheckIn': earliest,
                'CheckOut': latest,
                'GrossHours': round(gross_hours, 2),
            })
    if not records:
        return pd.DataFrame()
    return pd.DataFrame(records)


# ============================================================
# BIO MODE FUNCTIONS
# ============================================================
def process_monthly_matrix(file_bytes):
    df = pd.read_excel(io.BytesIO(file_bytes), header=0, dtype=str)
    header = df.columns.tolist()
    try:
        col_id = header.index('Employee ID') if 'Employee ID' in header else header.index('Employee_ID')
    except ValueError:
        st.error("Column 'Employee ID' not found.")
        return None, None
    try:
        col_name = header.index('First Name') if 'First Name' in header else header.index('First_Name')
    except ValueError:
        st.error("Column 'First Name' not found.")
        return None, None
    try:
        col_dept = header.index('Department')
    except ValueError:
        col_dept = -1
    try:
        col_regular = header.index('Regular')
    except ValueError:
        col_regular = len(header)
    start_date_col = col_dept + 1 if col_dept != -1 else 3
    end_date_col = col_regular if col_regular != len(header) else len(header)
    date_headers = header[start_date_col:end_date_col]
    date_objs = []
    for h in date_headers:
        try:
            dt = pd.to_datetime(h, format='%d-%m-%y')
            date_objs.append(dt)
        except:
            try:
                dt = pd.to_datetime(h)
                date_objs.append(dt)
            except:
                date_objs.append(None)
    records = []
    employee_names = {}
    for idx, row in df.iterrows():
        emp_id = str(row.iloc[col_id]).strip()
        if not emp_id or pd.isna(row.iloc[col_id]):
            continue
        name = str(row.iloc[col_name]).strip() if not pd.isna(row.iloc[col_name]) else ''
        employee_names[emp_id] = name
        for i, dt in enumerate(date_objs):
            if dt is None:
                continue
            cell_val = row.iloc[start_date_col + i]
            if pd.isna(cell_val) or cell_val == '':
                continue
            cell_str = str(cell_val).strip()
            if '-' in cell_str and ':' in cell_str:
                parts = cell_str.split('-')
                if len(parts) == 2:
                    start_str, end_str = parts[0].strip(), parts[1].strip()
                    try:
                        start_time = datetime.strptime(start_str, '%H:%M')
                        end_time = datetime.strptime(end_str, '%H:%M')
                        check_in = datetime.combine(dt.date(), start_time.time())
                        check_out = datetime.combine(dt.date(), end_time.time())
                        if check_out <= check_in:
                            check_out += timedelta(days=1)
                        gross_sec = (check_out - check_in).total_seconds()
                        gross_hours = gross_sec / 3600.0
                        records.append({
                            'EmployeeID': emp_id,
                            'Date': dt.date(),
                            'CheckIn': check_in,
                            'CheckOut': check_out,
                            'GrossHours': round(gross_hours, 2),
                        })
                    except:
                        pass
    return records, employee_names


# ============================================================
# COMMON ENRICHMENT & REPORT FUNCTIONS
# ============================================================
def assign_type(emp_id):
    return 'Permanent' if str(emp_id).startswith('100') else 'Temporary' if str(emp_id).startswith('200') else 'Unknown'

def get_normal_hours(emp_id, dt, enable_ramadan=False, ramadan_start=None, ramadan_end=None):
    if enable_ramadan and ramadan_start and ramadan_end:
        if ramadan_start <= dt <= ramadan_end:
            return RAMADAN_NORMAL_HOURS
    if str(emp_id).startswith('100'):
        return NORMAL_HOURS_PERMANENT
    elif str(emp_id).startswith('200'):
        return NORMAL_HOURS_TEMPORARY
    else:
        return 0

def enrich_attendance(df, break_start, break_end, break_duration_hours,
                      enable_friday_break, friday_break_start, friday_break_end, friday_break_duration,
                      enable_ramadan=False, ramadan_start=None, ramadan_end=None):
    if df.empty:
        return df

    df['EmployeeType'] = df['EmployeeID'].apply(assign_type)
    df['NormalHoursLimit'] = df.apply(
        lambda r: get_normal_hours(r['EmployeeID'], r['Date'], enable_ramadan, ramadan_start, ramadan_end), axis=1
    )
    df['Weekday'] = df['Date'].apply(lambda x: x.weekday())
    df['IsWeekend'] = df['Weekday'] == WEEKEND_DAY

    def apply_break(row):
        gross = row['GrossHours']
        if gross <= 6:
            return 0
        checkin = row['CheckIn']
        checkout = row['CheckOut']
        dt = checkin.date()
        if row['IsWeekend'] and enable_friday_break:
            b_start = friday_break_start
            b_end = friday_break_end
            break_dur = friday_break_duration
        else:
            b_start = break_start
            b_end = break_end
            break_dur = break_duration_hours
        break_start_dt = datetime.combine(dt, b_start)
        break_end_dt = datetime.combine(dt, b_end)
        if checkout <= break_start_dt or checkin >= break_end_dt:
            return 0
        return break_dur

    df['BreakHours'] = df.apply(apply_break, axis=1)
    df['NetHours'] = df['GrossHours'] - df['BreakHours']
    df['NetHours'] = df['NetHours'].round(2)

    def compute_hours(row):
        if row['IsWeekend']:
            normal = 0
            ot = row['NetHours']
        else:
            limit = row['NormalHoursLimit']
            normal = min(row['NetHours'], limit)
            ot = max(0, row['NetHours'] - limit)
        return pd.Series({'NormalHours': normal, 'Overtime': ot})

    df[['NormalHours', 'Overtime']] = df.apply(compute_hours, axis=1)

    def status(row):
        if row['IsWeekend']:
            return 'W' if row['NetHours'] == 0 else 'P'
        return 'P' if row['NetHours'] > 0 else 'A'
    df['Status'] = df.apply(status, axis=1)
    df['CheckInDisplay'] = df['CheckIn'].apply(lambda x: x.round('15min') if pd.notnull(x) else None)
    df['CheckOutDisplay'] = df['CheckOut'].apply(lambda x: x.round('15min') if pd.notnull(x) else None)
    return df

def generate_settlement_from_template(template_bytes, emp_id, attendance_df, employee_name, emp_type,
                                      start_date, end_date,
                                      break_start, break_end, break_duration_hours,
                                      enable_friday_break, friday_break_start, friday_break_end, friday_break_duration,
                                      enable_ramadan=False, ramadan_start=None, ramadan_end=None):
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb.active
    ws['F1'] = employee_name
    ws['F2'] = emp_id
    ws['F3'] = ''
    ws['C2'] = emp_type

    all_dates = pd.date_range(start=start_date, end=end_date, freq='D')
    emp_data = attendance_df[attendance_df['EmployeeID'] == str(emp_id).strip()]
    lookup = {row['Date']: row for _, row in emp_data.iterrows()}

    start_row = 7
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= start_row:
            ws.unmerge_cells(str(merged_range))

    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for i, dt in enumerate(all_dates):
        current_row = start_row + i
        date_val = dt.strftime('%Y-%m-%d')
        is_friday = dt.weekday() == WEEKEND_DAY
        row = lookup.get(dt.date())

        if is_friday:
            if row is not None and row['NetHours'] > 0:
                start_time = row['CheckInDisplay'].strftime('%H:%M:%S') if pd.notnull(row['CheckInDisplay']) else ''
                finish_time = row['CheckOutDisplay'].strftime('%H:%M:%S') if pd.notnull(row['CheckOutDisplay']) else ''
                normal_hours = 0
                ot_hours_display = round(row['NetHours'])
                ot_start = start_time
                ot_finish = finish_time
                status = 'P'
                remark = 'Friday (OT all day)'
            else:
                start_time = ''; finish_time = ''; normal_hours = 0; ot_hours_display = 0; ot_start = ''; ot_finish = ''; status = 'W'; remark = 'Week Off'
        else:
            if row is not None:
                start_time = row['CheckInDisplay'].strftime('%H:%M:%S') if pd.notnull(row['CheckInDisplay']) else ''
                finish_time = row['CheckOutDisplay'].strftime('%H:%M:%S') if pd.notnull(row['CheckOutDisplay']) else ''
                limit = get_normal_hours(emp_id, row['Date'], enable_ramadan, ramadan_start, ramadan_end)
                normal_hours = min(row['NetHours'], limit)
                ot_hours = max(0, row['NetHours'] - limit)
                ot_hours_display = round(ot_hours)
                normal_hours_display = round(normal_hours)
                if ot_hours_display > 0:
                    break_deducted = row['BreakHours'] > 0
                    if break_deducted:
                        ot_start_dt = row['CheckIn'] + timedelta(hours=limit + row['BreakHours'])
                    else:
                        ot_start_dt = row['CheckIn'] + timedelta(hours=limit)
                    ot_finish_dt = row['CheckOut']
                    ot_start = ot_start_dt.strftime('%H:%M')
                    ot_finish = ot_finish_dt.strftime('%H:%M:%S')
                else:
                    ot_start = ''; ot_finish = ''
                status = row['Status']; remark = ''
            else:
                start_time = ''; finish_time = ''; normal_hours = 0; normal_hours_display = 0; ot_hours_display = 0; ot_start = ''; ot_finish = ''; status = 'A'; remark = 'Absent'

        ws.cell(row=current_row, column=1, value=date_val)
        ws.cell(row=current_row, column=2, value=start_time)
        ws.cell(row=current_row, column=3, value=finish_time)
        ws.cell(row=current_row, column=4, value=normal_hours_display if 'normal_hours_display' in locals() else round(normal_hours))
        ws.cell(row=current_row, column=5, value=ot_start)
        ws.cell(row=current_row, column=6, value=ot_finish)
        ws.cell(row=current_row, column=7, value=ot_hours_display)
        ws.cell(row=current_row, column=8, value=status)
        ws.cell(row=current_row, column=9, value=remark)
        if is_friday:
            for col in range(1, 10):
                ws.cell(row=current_row, column=col).fill = grey_fill

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def generate_monthly_matrix(attendance_df, start_date, end_date):
    employees = attendance_df['EmployeeID'].unique()
    if len(employees) == 0:
        return pd.DataFrame(), pd.DatetimeIndex([])
    all_dates = pd.date_range(start=start_date, end=end_date, freq='D')
    rows = []
    for emp in employees:
        emp_att = attendance_df[attendance_df['EmployeeID'] == emp]
        name = emp_att['Name'].iloc[0] if 'Name' in emp_att.columns else emp
        row = {'Employee': emp, 'Name': name}
        for dt in all_dates:
            day = emp_att[emp_att['Date'] == dt.date()]
            if not day.empty:
                status = day.iloc[0]['Status']
                if dt.weekday() == WEEKEND_DAY:
                    if status == 'P':
                        att_val = '1'; ot_val = round(day.iloc[0]['NetHours'])
                    else:
                        att_val = 'W'; ot_val = 0
                else:
                    att_val = '1' if status == 'P' else 'A'
                    ot_val = round(day.iloc[0]['Overtime'])
            else:
                att_val = 'W' if dt.weekday() == WEEKEND_DAY else 'A'
                ot_val = 0
            row[f'{dt.strftime("%Y-%m-%d")}_Att'] = att_val
            row[f'{dt.strftime("%Y-%m-%d")}_OT'] = ot_val
        att_cols = [col for col in row.keys() if col.endswith('_Att')]
        ot_cols = [col for col in row.keys() if col.endswith('_OT')]
        total_present = sum(1 for col in att_cols if row[col] == '1')
        total_absent = sum(1 for col in att_cols if row[col] == 'A')
        total_weekends = sum(1 for col in att_cols if row[col] == 'W')
        total_ot = sum(row[col] for col in ot_cols if isinstance(row[col], (int, float)))
        row['Total_Present'] = total_present
        row['Total_Absent'] = total_absent
        row['Total_Weekends'] = total_weekends
        row['Total_OT'] = total_ot
        rows.append(row)
    df = pd.DataFrame(rows)
    df.set_index(['Employee', 'Name'], inplace=True)
    return df, all_dates

def export_monthly_matrix(df, all_dates, start_date, end_date):
    if df.empty or len(all_dates) == 0:
        return None
    output = io.BytesIO()
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Monthly', index=True)
            wb = writer.book
            ws = wb['Monthly']
            ws.sheet_state = 'visible'
            ws.insert_rows(1)
            start_col = 3
            for idx, dt in enumerate(all_dates):
                col = start_col + idx * 2
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
                ws.cell(row=1, column=col, value=dt.strftime('%Y-%m-%d'))
                ws.cell(row=2, column=col, value='Att')
                ws.cell(row=2, column=col+1, value='OT')
            grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            header_font = Font(color="000000", bold=True)
            alt_fill1 = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
            alt_fill2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            for idx, dt in enumerate(all_dates):
                if dt.weekday() == WEEKEND_DAY:
                    col = start_col + idx * 2
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=col).fill = grey_fill
                        ws.cell(row=row, column=col+1).fill = grey_fill
            for row_num in [1, 2]:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            emp_rows = []
            current_emp = None
            for row_idx in range(3, ws.max_row + 1):
                emp_val = ws.cell(row=row_idx, column=1).value
                if emp_val and emp_val != current_emp:
                    current_emp = emp_val
                    fill = alt_fill1 if (len(emp_rows) % 2 == 0) else alt_fill2
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=c).fill = fill
                    emp_rows.append(current_emp)
                else:
                    if emp_rows:
                        fill = alt_fill1 if (len(emp_rows) % 2 == 1) else alt_fill2
                        for c in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=c).fill = fill
            center_alignment = Alignment(horizontal='center', vertical='center')
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = center_alignment
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                for row in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        return output.getvalue()
    except Exception as e:
        st.error(f"Error generating Excel: {e}")
        return None


# ============================================================
# APP RENDER FUNCTIONS
# ============================================================
def render_machine_app():
    st.title("📊 Machine Attendance System")
    st.markdown("Upload raw logs and settlement template – everything else is automatic.")

    st.sidebar.header("⚙️ Settings")
    global NORMAL_HOURS_PERMANENT, NORMAL_HOURS_TEMPORARY
    NORMAL_HOURS_PERMANENT = st.sidebar.number_input("Permanent Normal Hours (net)", value=8, step=1)
    NORMAL_HOURS_TEMPORARY = st.sidebar.number_input("Temporary Normal Hours (net)", value=10, step=1)

    st.sidebar.subheader("🕒 Normal Break (Sat–Thu)")
    break_start = st.sidebar.time_input("Break Start", value=time(12, 0))
    break_end = st.sidebar.time_input("Break End", value=time(13, 0))
    if break_start >= break_end:
        st.sidebar.error("Break start must be before end.")
    else:
        break_delta = datetime.combine(date.today(), break_end) - datetime.combine(date.today(), break_start)
        break_duration_hours = break_delta.total_seconds() / 3600.0

    st.sidebar.subheader("🕒 Friday Break (Optional)")
    enable_friday_break = st.sidebar.checkbox("Enable break on Fridays", value=False)
    if enable_friday_break:
        friday_break_start = st.sidebar.time_input("Friday Break Start", value=time(12, 0))
        friday_break_end = st.sidebar.time_input("Friday Break End", value=time(13, 0))
        if friday_break_start >= friday_break_end:
            st.sidebar.error("Friday break start must be before end.")
        else:
            friday_delta = datetime.combine(date.today(), friday_break_end) - datetime.combine(date.today(), friday_break_start)
            friday_break_duration = friday_delta.total_seconds() / 3600.0
    else:
        friday_break_start = time(12,0); friday_break_end = time(13,0); friday_break_duration = 0

    st.sidebar.subheader("🌙 Ramadan Period")
    enable_ramadan = st.sidebar.checkbox("Enable Ramadan", value=False)
    if enable_ramadan:
        ramadan_start = st.sidebar.date_input("Ramadan Start Date", value=date(2026, 3, 1))
        ramadan_end = st.sidebar.date_input("Ramadan End Date", value=date(2026, 3, 30))
        if ramadan_start > ramadan_end:
            st.sidebar.error("Ramadan start must be before end.")
    else:
        ramadan_start = None; ramadan_end = None

    st.sidebar.header("📁 Upload Files")
    logs_file = st.sidebar.file_uploader("Raw Logs (Excel with ID, DATE)", type=["xlsx"])
    template_file = st.sidebar.file_uploader("Settlement Template (Final Settlement.xlsx)", type=["xlsx"])

    st.sidebar.markdown("---")
    st.sidebar.write("👤 Ahmed Shawky")
    st.sidebar.write("📞 +201095214911")
    st.sidebar.write("✉️ ahmedshawkyqz@gmail.com")
    st.sidebar.markdown("---")

    if logs_file is None:
        st.info("Please upload the raw logs to begin.")
        return

    with st.spinner("Loading and processing logs..."):
        logs_df = pd.read_excel(logs_file, dtype={'ID': str})
        if 'ID' not in logs_df.columns or 'DATE' not in logs_df.columns:
            st.error("Logs file must contain columns 'ID' and 'DATE'.")
            return
        daily = process_logs(logs_df)
        if daily.empty:
            st.error("No valid attendance records found.")
            return
        daily = enrich_attendance(daily,
                                  break_start, break_end, break_duration_hours,
                                  enable_friday_break, friday_break_start, friday_break_end, friday_break_duration,
                                  enable_ramadan, ramadan_start, ramadan_end)
        daily['Name'] = daily['EmployeeID']

    st.success(f"Processed {len(daily)} daily records for {daily['EmployeeID'].nunique()} employees.")

    st.sidebar.subheader("📊 Data Info")
    st.sidebar.write(f"**Employees:** {daily['EmployeeID'].nunique()}")
    min_date = daily['Date'].min()
    max_date = daily['Date'].max()
    st.sidebar.write(f"**Date range:** {min_date} to {max_date}")

    tab1, tab2, tab3, tab4 = st.tabs(["📋 Preview", "🧾 Settlements", "📅 Monthly Report", "✏️ Edit"])

    with tab1:
        st.subheader("All Daily Records (scrollable)")
        st.dataframe(daily[['EmployeeID', 'Name', 'Date', 'CheckInDisplay', 'CheckOutDisplay',
                            'GrossHours', 'BreakHours', 'NetHours', 'NormalHours', 'Overtime', 'Status']])

    with tab2:
        if template_file is None:
            st.warning("Please upload the settlement template file.")
        else:
            emp_list = daily['EmployeeID'].unique().tolist()
            selected = st.selectbox("Select Employee ID", emp_list)
            st.subheader("Select Date Range")
            range_option = st.radio("Choose range type:", ["Month", "Custom Range"], key="range_settlement")
            if range_option == "Month":
                year = st.number_input("Year", min_value=2020, max_value=2030, value=2026, key="year_settle")
                month = st.selectbox("Month", range(1,13), format_func=lambda x: datetime(year, x, 1).strftime("%B"), key="month_settle")
                start_date = date(year, month, 1)
                if month == 12:
                    end_date = date(year+1, 1, 1) - timedelta(days=1)
                else:
                    end_date = date(year, month+1, 1) - timedelta(days=1)
            else:
                col1, col2 = st.columns(2)
                with col1:
                    start_date = st.date_input("Start Date", value=date(2026, 6, 1), key="start_settle")
                with col2:
                    end_date = st.date_input("End Date", value=date(2026, 9, 30), key="end_settle")
                if start_date > end_date:
                    st.error("Start date must be before end date.")
                    return
            emp_data_in_range = daily[(daily['EmployeeID'] == selected) & (daily['Date'] >= start_date) & (daily['Date'] <= end_date)]
            if emp_data_in_range.empty:
                st.warning(f"No attendance records found for employee {selected} in the selected date range.")
            else:
                st.success(f"Found {len(emp_data_in_range)} records for this employee in the range.")
            if st.button("📄 Generate Settlement", key="gen_settle"):
                emp_name = daily[daily['EmployeeID']==selected]['Name'].iloc[0]
                emp_type = assign_type(selected)
                template_bytes = template_file.getvalue()
                excel_data = generate_settlement_from_template(
                    template_bytes, selected, daily, emp_name, emp_type, start_date, end_date,
                    break_start, break_end, break_duration_hours,
                    enable_friday_break, friday_break_start, friday_break_end, friday_break_duration,
                    enable_ramadan, ramadan_start, ramadan_end
                )
                if excel_data:
                    st.download_button(label="📥 Download Settlement", data=excel_data, file_name=f"Settlement_{selected}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    st.success("Settlement generated!")
                else:
                    st.error("Failed to generate settlement.")
            if st.button("📦 Generate All Settlements (ZIP)", key="gen_all_settle"):
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                    for eid in emp_list:
                        emp_name = daily[daily['EmployeeID']==eid]['Name'].iloc[0]
                        emp_type = assign_type(eid)
                        template_bytes = template_file.getvalue()
                        excel_data = generate_settlement_from_template(
                            template_bytes, eid, daily, emp_name, emp_type, start_date, end_date,
                            break_start, break_end, break_duration_hours,
                            enable_friday_break, friday_break_start, friday_break_end, friday_break_duration,
                            enable_ramadan, ramadan_start, ramadan_end
                        )
                        if excel_data:
                            zip_file.writestr(f"Settlement_{eid}.xlsx", excel_data)
                zip_buffer.seek(0)
                st.download_button(label="📥 Download All Settlements (ZIP)", data=zip_buffer, file_name="All_Settlements.zip", mime="application/zip")
                st.success("All settlements zipped!")

    with tab3:
        st.subheader("📅 Monthly Attendance Matrix (two columns per day: Att / OT)")
        col1, col2 = st.columns(2)
        with col1:
            start_date = st.date_input("Start Date", value=date(2026, 6, 1), key="start_monthly")
        with col2:
            end_date = st.date_input("End Date", value=date(2026, 9, 30), key="end_monthly")
        if start_date > end_date:
            st.error("Start date must be before end date.")
        else:
            if st.button("Generate Monthly Report", key="gen_monthly"):
                with st.spinner("Generating..."):
                    matrix_df, all_dates = generate_monthly_matrix(daily, start_date, end_date)
                    if matrix_df.empty:
                        st.warning("No data available for the selected date range.")
                    else:
                        st.dataframe(matrix_df)
                        excel_data = export_monthly_matrix(matrix_df, all_dates, start_date, end_date)
                        if excel_data:
                            st.download_button(label="📥 Download Monthly Report", data=excel_data, file_name=f"Monthly_Attendance_{start_date.strftime('%b%Y')}_{end_date.strftime('%b%Y')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                            st.success("Monthly report ready!")
                        else:
                            st.error("Failed to generate the Excel file. Please try again.")

    with tab4:
        st.subheader("✏️ Edit Attendance Data (changes reflect in exports)")
        edited = st.data_editor(daily, num_rows="dynamic", key="data_editor")
        if st.button("Update Attendance", key="update_data"):
            st.session_state['daily'] = edited
            st.success("Updated!")


def render_bio_app():
    st.title("📊 Bio Attendance System (Monthly Matrix)")
    st.markdown("Upload monthly attendance matrix and settlement template – everything else is automatic.")

    st.sidebar.header("⚙️ Settings")
    global NORMAL_HOURS_PERMANENT, NORMAL_HOURS_TEMPORARY
    NORMAL_HOURS_PERMANENT = st.sidebar.number_input("Permanent Normal Hours (net)", value=8, step=1)
    NORMAL_HOURS_TEMPORARY = st.sidebar.number_input("Temporary Normal Hours (net)", value=10, step=1)

    st.sidebar.subheader("🕒 Normal Break (Sat–Thu)")
    break_start = st.sidebar.time_input("Break Start", value=time(12, 0))
    break_end = st.sidebar.time_input("Break End", value=time(13, 0))
    if break_start >= break_end:
        st.sidebar.error("Break start must be before end.")
    else:
        break_delta = datetime.combine(date.today(), break_end) - datetime.combine(date.today(), break_start)
        break_duration_hours = break_delta.total_seconds() / 3600.0

    st.sidebar.subheader("🕒 Friday Break (Optional)")
    enable_friday_break = st.sidebar.checkbox("Enable break on Fridays", value=False)
    if enable_friday_break:
        friday_break_start = st.sidebar.time_input("Friday Break Start", value=time(12, 0))
        friday_break_end = st.sidebar.time_input("Friday Break End", value=time(13, 0))
        if friday_break_start >= friday_break_end:
            st.sidebar.error("Friday break start must be before end.")
        else:
            friday_delta = datetime.combine(date.today(), friday_break_end) - datetime.combine(date.today(), friday_break_start)
            friday_break_duration = friday_delta.total_seconds() / 3600.0
    else:
        friday_break_start = time(12,0); friday_break_end = time(13,0); friday_break_duration = 0

    st.sidebar.subheader("🌙 Ramadan Period")
    enable_ramadan = st.sidebar.checkbox("Enable Ramadan", value=False)
    if enable_ramadan:
        ramadan_start = st.sidebar.date_input("Ramadan Start Date", value=date(2026, 3, 1))
        ramadan_end = st.sidebar.date_input("Ramadan End Date", value=date(2026, 3, 30))
        if ramadan_start > ramadan_end:
            st.sidebar.error("Ramadan start must be before end.")
    else:
        ramadan_start = None; ramadan_end = None

    st.sidebar.header("📁 Upload Files")
    logs_file = st.sidebar.file_uploader("Monthly Attendance Matrix (Excel with ID, Name, Dates)", type=["xlsx"])
    template_file = st.sidebar.file_uploader("Settlement Template (Final Settlement.xlsx)", type=["xlsx"])

    st.sidebar.markdown("---")
    st.sidebar.write("👤 Ahmed Shawky")
    st.sidebar.write("📞 +201095214911")
    st.sidebar.write("✉️ ahmedshawkyqz@gmail.com")
    st.sidebar.markdown("---")

    if logs_file is None:
        st.info("Please upload the monthly attendance matrix to begin.")
        return

    with st.spinner("Processing attendance matrix..."):
        file_bytes = logs_file.getvalue()
        records, employee_names = process_monthly_matrix(file_bytes)
        if records is None:
            st.error("Failed to parse the file. Please check the format.")
            return
        if not records:
            st.error("No attendance records found in the file.")
            return
        daily = pd.DataFrame(records)
        daily['Name'] = daily['EmployeeID'].map(employee_names)
        daily['Name'] = daily['Name'].fillna(daily['EmployeeID'])

        daily = enrich_attendance(daily,
                                  break_start, break_end, break_duration_hours,
                                  enable_friday_break, friday_break_start, friday_break_end, friday_break_duration,
                                  enable_ramadan, ramadan_start, ramadan_end)

    st.success(f"Processed {len(daily)} daily records for {daily['EmployeeID'].nunique()} employees.")

    st.sidebar.subheader("📊 Data Info")
    st.sidebar.write(f"**Employees:** {daily['EmployeeID'].nunique()}")
    min_date = daily['Date'].min()
    max_date = daily['Date'].max()
    st.sidebar.write(f"**Date range:** {min_date} to {max_date}")

    tab1, tab2, tab3, tab4 = st.tabs(["📋 Preview", "🧾 Settlements", "📅 Monthly Report", "✏️ Edit"])

    with tab1:
        st.subheader("All Daily Records (scrollable)")
        st.dataframe(daily[['EmployeeID', 'Name', 'Date', 'CheckInDisplay', 'CheckOutDisplay',
                            'GrossHours', 'BreakHours', 'NetHours', 'NormalHours', 'Overtime', 'Status']])

    with tab2:
        if template_file is None:
            st.warning("Please upload the settlement template file.")
        else:
            emp_list = daily['EmployeeID'].unique().tolist()
            selected = st.selectbox("Select Employee ID", emp_list)
            st.subheader("Select Date Range")
            range_option = st.radio("Choose range type:", ["Month", "Custom Range"], key="range_settlement")
            if range_option == "Month":
                year = st.number_input("Year", min_value=2020, max_value=2030, value=2026, key="year_settle")
                month = st.selectbox("Month", range(1,13), format_func=lambda x: datetime(year, x, 1).strftime("%B"), key="month_settle")
                start_date = date(year, month, 1)
                if month == 12:
                    end_date = date(year+1, 1, 1) - timedelta(days=1)
                else:
                    end_date = date(year, month+1, 1) - timedelta(days=1)
            else:
                col1, col2 = st.columns(2)
                with col1:
                    start_date = st.date_input("Start Date", value=date(2026, 6, 1), key="start_settle")
                with col2:
                    end_date = st.date_input("End Date", value=date(2026, 9, 30), key="end_settle")
                if start_date > end_date:
                    st.error("Start date must be before end date.")
                    return
            emp_data_in_range = daily[(daily['EmployeeID'] == selected) & (daily['Date'] >= start_date) & (daily['Date'] <= end_date)]
            if emp_data_in_range.empty:
                st.warning(f"No attendance records found for employee {selected} in the selected date range.")
            else:
                st.success(f"Found {len(emp_data_in_range)} records for this employee in the range.")
            if st.button("📄 Generate Settlement", key="gen_settle"):
                emp_name = daily[daily['EmployeeID']==selected]['Name'].iloc[0]
                emp_type = assign_type(selected)
                template_bytes = template_file.getvalue()
                excel_data = generate_settlement_from_template(
                    template_bytes, selected, daily, emp_name, emp_type, start_date, end_date,
                    break_start, break_end, break_duration_hours,
                    enable_friday_break, friday_break_start, friday_break_end, friday_break_duration,
                    enable_ramadan, ramadan_start, ramadan_end
                )
                if excel_data:
                    st.download_button(label="📥 Download Settlement", data=excel_data, file_name=f"Settlement_{selected}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    st.success("Settlement generated!")
                else:
                    st.error("Failed to generate settlement.")
            if st.button("📦 Generate All Settlements (ZIP)", key="gen_all_settle"):
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                    for eid in emp_list:
                        emp_name = daily[daily['EmployeeID']==eid]['Name'].iloc[0]
                        emp_type = assign_type(eid)
                        template_bytes = template_file.getvalue()
                        excel_data = generate_settlement_from_template(
                            template_bytes, eid, daily, emp_name, emp_type, start_date, end_date,
                            break_start, break_end, break_duration_hours,
                            enable_friday_break, friday_break_start, friday_break_end, friday_break_duration,
                            enable_ramadan, ramadan_start, ramadan_end
                        )
                        if excel_data:
                            zip_file.writestr(f"Settlement_{eid}.xlsx", excel_data)
                zip_buffer.seek(0)
                st.download_button(label="📥 Download All Settlements (ZIP)", data=zip_buffer, file_name="All_Settlements.zip", mime="application/zip")
                st.success("All settlements zipped!")

    with tab3:
        st.subheader("📅 Monthly Attendance Matrix (two columns per day: Att / OT)")
        col1, col2 = st.columns(2)
        with col1:
            start_date = st.date_input("Start Date", value=date(2026, 6, 1), key="start_monthly")
        with col2:
            end_date = st.date_input("End Date", value=date(2026, 9, 30), key="end_monthly")
        if start_date > end_date:
            st.error("Start date must be before end date.")
        else:
            if st.button("Generate Monthly Report", key="gen_monthly"):
                with st.spinner("Generating..."):
                    matrix_df, all_dates = generate_monthly_matrix(daily, start_date, end_date)
                    if matrix_df.empty:
                        st.warning("No data available for the selected date range.")
                    else:
                        st.dataframe(matrix_df)
                        excel_data = export_monthly_matrix(matrix_df, all_dates, start_date, end_date)
                        if excel_data:
                            st.download_button(label="📥 Download Monthly Report", data=excel_data, file_name=f"Monthly_Attendance_{start_date.strftime('%b%Y')}_{end_date.strftime('%b%Y')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                            st.success("Monthly report ready!")
                        else:
                            st.error("Failed to generate the Excel file. Please try again.")

    with tab4:
        st.subheader("✏️ Edit Attendance Data (changes reflect in exports)")
        edited = st.data_editor(daily, num_rows="dynamic", key="data_editor")
        if st.button("Update Attendance", key="update_data"):
            st.session_state['daily'] = edited
            st.success("Updated!")


# ============================================================
# MAIN PROGRAM
# ============================================================
def main():
    st.set_page_config(
        page_title="Ultimate Attendance System",
        layout="wide",
        initial_sidebar_state="collapsed"
    )

    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
    if "app_mode" not in st.session_state:
        st.session_state.app_mode = None

    if not st.session_state.logged_in:
        render_login()
        return

    if st.session_state.app_mode is None:
        render_mode_selection()
        return

    st.set_page_config(initial_sidebar_state="expanded")

    with st.sidebar:
        if st.button("🔄 Switch Mode"):
            st.session_state.app_mode = None
            st.rerun()

    if st.session_state.app_mode == "machine":
        render_machine_app()
    elif st.session_state.app_mode == "bio":
        render_bio_app()
    else:
        st.error("Invalid mode selected.")

if __name__ == "__main__":
    main()
